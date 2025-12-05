import streamlit as st
import pandas as pd
import json
import math
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------------- 数据加载模块 ----------------------
def load_pricing_rules(file_path):
    """加载加价规则"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            rules = json.load(f)
        
        preprocessed = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
        for mill, mill_data in rules.items():
            for price_type, price_data in mill_data.items():
                for steel_type, type_data in price_data.items():
                    base_spec = type_data['base_spec']
                    preprocessed[mill][steel_type][price_type]['base_spec'] = base_spec
                    for spec, addition in type_data['additions'].items():
                        preprocessed[mill][steel_type][price_type][spec] = addition
        
        return preprocessed
    except FileNotFoundError:
        st.error(f"错误：加价规则文件 {file_path} 未找到")
        return None
    except json.JSONDecodeError:
        st.error(f"错误：加价规则文件 {file_path} 格式错误")
        return None

def load_steel_weights(file_path):
    """加载钢厂件重数据"""
    try:
        df = pd.read_csv(file_path)
        weight_dict = {}
        for _, row in df.iterrows():
            mill = row['钢厂品牌']
            spec = row['规格型号']
            length = row['长度'] if not pd.isna(row['长度']) else ''
            weight = row['重量（件）']
            
            if mill not in weight_dict:
                weight_dict[mill] = {}
            if spec not in weight_dict[mill]:
                weight_dict[mill][spec] = {}
            weight_dict[mill][spec][length] = weight
        return weight_dict
    except FileNotFoundError:
        st.error(f"错误：件重数据文件 {file_path} 未找到")
        return None
    except Exception as e:
        st.error(f"加载件重数据失败：{str(e)}")
        return None

# ---------------------- 价格计算模块 ----------------------
def calculate_price_diff(steel_mill, steel_type, spec, length, base_net_price, base_arrival_price, pricing_rules, enable_12m_addition=True):
    """计算价差，添加12m加价规则开关"""
    try:
        price_rules = pricing_rules[steel_mill][steel_type]
        if '网价' not in price_rules or '到货价' not in price_rules:
            return 0, base_net_price, base_arrival_price
        
        net_addition = price_rules['网价'].get(spec, 0)
        arrival_addition = price_rules['到货价'].get(spec, 0)
        
        net_price = base_net_price + net_addition
        arrival_price = base_arrival_price + arrival_addition
        
        # 仅当启用加价且满足条件时才应用30元加价
        if (enable_12m_addition and 
            steel_mill in ['徐钢', '河南闽源', '中新'] and 
            length == '12m' and 
            spec in ['HRB400E12', 'HRB400E14', 'HRB400E16', 'HRB400E18', 
                     'HRB400E20', 'HRB400E22', 'HRB400E25']):
            net_price += 30
        
        return net_price - arrival_price, net_price, arrival_price
    except KeyError:
        return 0, base_net_price, base_arrival_price

# ---------------------- 辅助计算模块 ----------------------
def calculate_ship_pieces(tonnage, weight, tolerance=1.0):
    """
    计算发货件数，使实际吨位与计划吨位的偏差不超过允许范围
    
    参数:
    tonnage: 计划吨位
    weight: 每件重量
    tolerance: 允许的最大偏差（吨）
    
    返回:
    ship_pieces: 发货件数
    ship_weight: 实际发货吨位
    """
    if weight <= 0 or tonnage <= 0:
        return 0, 0
    
    base_pieces = int(tonnage / weight)
    base_weight = base_pieces * weight
    diff = tonnage - base_weight
    
    if base_pieces == 0:
        return 1, weight
    if diff > tolerance:
        return base_pieces + 1, (base_pieces + 1) * weight
    return base_pieces, base_weight

# ---------------------- 价差表生成模块 ----------------------
def get_weight(steel_mill, spec, length, weight_dict):
    """获取件重"""
    try:
        # 盘螺长度为空字符串
        if spec in ['HRB400E6', 'HRB400E8', 'HRB400E10']:
            return weight_dict[steel_mill][spec]['']
        return weight_dict[steel_mill][spec][length]
    except KeyError:
        return None

def generate_manual_pricing_table(shipment_plan, available_specs, weight_dict, pricing_rules, base_prices_dict, 
                                 enable_12m_addition=True, tonnage_tolerance=1.0):
    """生成人工发货价差表，包含优化的件数计算逻辑和规格汇总行"""
    all_candidates = []
    available_specs_set = {mill: {spec: set(lengths) for spec, lengths in specs.items()} 
                          for mill, specs in available_specs.items()}
    
    # 收集所有可发规格，用于检查无库存情况
    all_available_specs = set()
    for mill_specs in available_specs.values():
        all_available_specs.update(mill_specs.keys())
    
    # 1. 生成规格汇总行（楼号为空）
    summary_rows = []
    for spec in all_available_specs:
        max_price_diff = -1
        best_record = None
        
        # 查找该规格的最高价差记录
        for steel_mill in available_specs:
            if spec not in available_specs[steel_mill]:
                continue
                
            steel_type = '盘螺' if spec in ['HRB400E6', 'HRB400E8', 'HRB400E10'] else '螺纹钢'
            if steel_mill not in base_prices_dict or steel_type not in base_prices_dict[steel_mill]:
                continue
                
            lengths = available_specs_set[steel_mill].get(spec, set())
            base_prices = base_prices_dict[steel_mill][steel_type]
            base_net_price = base_prices.get('网价', 0)
            base_arrival_price = base_prices.get('到货价', 0)
            
            for length in lengths:
                price_diff, net_price, arrival_price = calculate_price_diff(
                    steel_mill, steel_type, spec, length, base_net_price, base_arrival_price,
                    pricing_rules, enable_12m_addition=enable_12m_addition
                )
                
                if price_diff > max_price_diff:
                    max_price_diff = price_diff
                    weight = get_weight(steel_mill, spec, length, weight_dict) or 0
                    
                    best_record = {
                        '楼号': '',  # 楼号为空
                        '规格': spec,
                        '规格排序': -1,  # 确保汇总行排在最前面
                        '长度': length,
                        '钢厂': steel_mill,
                        '网价(元/吨)': net_price,
                        '到货价(元/吨)': arrival_price,
                        '价差（元/吨）': price_diff,
                        '件重(吨)': round(weight, 3) if weight else 0,
                        '计划吨位': 0,
                        '发货件数': 0,
                        '发货吨位': 0,
                        '总利润(元)': 0,
                        'is_max_diff': True,  # 标记为最高价差行
                        'is_out_of_stock': False,
                        'is_summary': True  # 标记为汇总行
                    }
        
        if best_record:
            summary_rows.append(best_record)
    
    # 2. 生成原有数据行
    for building, specs in shipment_plan.items():
        for spec, tonnage in specs.items():
            # 检查该规格是否有可发钢厂
            if spec not in all_available_specs:
                # 无库存规格，添加特殊记录
                all_candidates.append({
                    '楼号': building,
                    '规格': spec,
                    '规格排序': int(''.join(filter(str.isdigit, spec))) if any(c.isdigit() for c in spec) else 0,
                    '长度': '',
                    '钢厂': '无库存',  # 标记无库存
                    '网价(元/吨)': 0,
                    '到货价(元/吨)': 0,
                    '价差（元/吨）': 0,
                    '件重(吨)': 0,
                    '计划吨位': round(tonnage, 2),
                    '发货件数': 0,
                    '发货吨位': 0,
                    '总利润(元)': 0,
                    'is_max_diff': False,
                    'is_out_of_stock': True,  # 添加无库存标记
                    'is_summary': False
                })
                continue
            
            # 有库存规格，但可能没有有效价差的情况
            steel_type = '盘螺' if spec in ['HRB400E6', 'HRB400E8', 'HRB400E10'] else '螺纹钢'
            candidate_mills = [mill for mill in available_specs if spec in available_specs[mill]]
            valid_records = 0  # 记录有效记录数量
            
            for steel_mill in candidate_mills:
                if steel_mill not in base_prices_dict or steel_type not in base_prices_dict[steel_mill]:
                    continue
                    
                lengths = available_specs_set[steel_mill].get(spec, set())
                if not lengths:
                    continue
                    
                base_prices = base_prices_dict[steel_mill][steel_type]
                base_net_price = base_prices.get('网价', 0)
                base_arrival_price = base_prices.get('到货价', 0)
                
                for length in lengths:
                    price_diff, net_price, arrival_price = calculate_price_diff(
                        steel_mill, steel_type, spec, length, base_net_price, base_arrival_price,
                        pricing_rules, enable_12m_addition=enable_12m_addition
                    )
                    
                    if price_diff <= 0:
                        continue
                    
                    weight = get_weight(steel_mill, spec, length, weight_dict)
                    if weight is None or weight <= 0:
                        continue
                    
                    # 使用新的件数计算函数
                    ship_pieces, ship_weight = calculate_ship_pieces(
                        tonnage, weight, tolerance=tonnage_tolerance
                    )
                    
                    spec_num = int(''.join(filter(str.isdigit, spec))) if any(c.isdigit() for c in spec) else 0
                    
                    # 修改总利润计算公式为：发货吨位 × 价差
                    total_profit = round(ship_weight * price_diff, 2)
                    
                    all_candidates.append({
                        '楼号': building,
                        '规格': spec,
                        '规格排序': spec_num,
                        '长度': length,
                        '钢厂': steel_mill,
                        '网价(元/吨)': net_price,
                        '到货价(元/吨)': arrival_price,
                        '价差（元/吨）': price_diff,
                        '件重(吨)': round(weight, 3),
                        '计划吨位': round(tonnage, 2),
                        '发货件数': ship_pieces,
                        '发货吨位': round(ship_weight, 3),  # 保留3位小数
                        '总利润(元)': total_profit,  # 使用新公式计算的总利润
                        'is_max_diff': False,
                        'is_out_of_stock': False,
                        'is_summary': False
                    })
                    valid_records += 1
            
            # 如果有可发规格但没有有效价差记录（如所有价差<=0）
            if valid_records == 0:
                all_candidates.append({
                    '楼号': building,
                    '规格': spec,
                    '规格排序': int(''.join(filter(str.isdigit, spec))) if any(c.isdigit() for c in spec) else 0,
                    '长度': '',
                    '钢厂': '无有效价差',  # 标记无有效价差
                    '网价(元/吨)': 0,
                    '到货价(元/吨)': 0,
                    '价差（元/吨）': 0,
                    '件重(吨)': 0,
                    '计划吨位': round(tonnage, 2),
                    '发货件数': 0,
                    '发货吨位': 0,
                    '总利润(元)': 0,
                    'is_max_diff': False,
                    'is_out_of_stock': True,  # 视为无库存
                    'is_summary': False
                })
    
    # 合并汇总行和数据行
    all_rows = summary_rows + all_candidates
    
    if not all_rows:
        return pd.DataFrame()
    
    df = pd.DataFrame(all_rows)
    
    # 排序：汇总行（规格排序=-1）排在最前面，然后按楼号、规格排序、价差排序
    df = df.sort_values(
        ['规格排序', '楼号', '价差（元/吨）'], 
        ascending=[True, True, False]
    )
    
    # 按楼号和规格分组，标记每组中价差最高的行（仅用于数据处理）
    df['is_max_diff'] = df.groupby(['楼号', '规格'])['价差（元/吨）'].transform(lambda x: x == x.max() if not x.empty else False)
    
    return df

def format_manual_table(df):
    """格式化人工发货价差表，移除高亮显示"""
    if df.empty:
        return df
    
    # 选择要显示的列
    display_df = df[['楼号', '规格', '长度', '钢厂', '网价(元/吨)', '到货价(元/吨)', '价差（元/吨）',
                    '件重(吨)', '计划吨位', '发货件数', '发货吨位', '总利润(元)']]
    
    # 创建样式器（仅保留基础格式）
    styler = display_df.style
    
    # 基础格式化
    styler = styler.set_properties(**{'text-align': 'center'})
    styler = styler.set_table_styles([
        {'selector': 'th', 'props': [('background-color', '#e3f2fd'), ('font-weight', 'bold'), ('text-align', 'center')]}
    ])
    
    return styler

def format_excel_with_highlight(df):
    """生成格式化的Excel文件，为规格汇总行添加高亮显示"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 将数据写入Excel（不包含辅助列）
        export_df = df[[col for col in df.columns if col not in ['规格排序', 'is_max_diff', 'is_out_of_stock', 'is_summary']]]
        export_df.to_excel(writer, index=False, sheet_name='发货依据')
        ws = writer.sheets['发货依据']
        
        # 设置表头样式
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:  # 表头在第2行（索引1）
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 高亮规格汇总行（楼号为空且is_summary=True）
        summary_mask = df['is_summary']
        summary_indices = df[summary_mask].index + 2  # +2是因为Excel行号从1开始且表头占1行
        
        # 设置高亮样式（黄色背景）
        highlight_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        
        for row_num in summary_indices:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = highlight_fill
        
        # 调整列宽
        column_widths = {
            '楼号': 8, '规格': 12, '钢厂': 10,
            '网价(元/吨)': 12, '到货价(元/吨)': 12, '价差（元/吨）': 12,
            '长度': 8, '件重(吨)': 10, '计划吨位': 10,
            '发货件数': 10, '发货吨位': 10, '总利润(元)': 12
        }
        for col_name, width in column_widths.items():
            if col_name in export_df.columns:
                col_idx = export_df.columns.get_loc(col_name)
                col_letter = chr(col_idx + 65)  # A=65
                ws.column_dimensions[col_letter].width = width
    
    return output.getvalue()

# ---------------------- 输入处理模块 ----------------------
def load_daily_base_prices(uploaded_file):
    """加载每日基价CSV"""
    try:
        df = pd.read_csv(uploaded_file)
        required_columns = ['钢厂', '钢筋类型', '网价基价', '到货价基价']
        if not all(col in df.columns for col in required_columns):
            st.error(f"基价文件缺少必要列，请确保包含：{required_columns}")
            return None
        
        df = df[df['网价基价'] > 0]
        df = df[df['到货价基价'] > 0]
        
        base_prices_dict = defaultdict(lambda: defaultdict(dict))
        for _, row in df.iterrows():
            mill = row['钢厂']
            steel_type = row['钢筋类型']
            if steel_type not in ['螺纹钢', '盘螺']:
                continue
            
            base_prices_dict[mill][steel_type] = {
                '网价': row['网价基价'],
                '到货价': row['到货价基价']
            }
        return base_prices_dict
    except Exception as e:
        st.error(f"加载基价文件失败：{str(e)}")
        return None

def load_available_specs(uploaded_file):
    """加载每日可发规格CSV，支持是否可发列和钢厂间空行"""
    try:
        # 读取CSV时保留空行，以便处理钢厂间空行分隔
        df = pd.read_csv(uploaded_file, skip_blank_lines=False)
        
        # 检查必要列
        required_columns = ['钢厂', '规格型号', '是否可发']
        if not all(col in df.columns for col in required_columns):
            st.error(f"可发规格文件缺少必要列，请确保包含：{required_columns}")
            return None
        
        # 添加长度列（如果不存在）
        if '长度' not in df.columns:
            df['长度'] = ''
        
        # 处理空行和钢厂填充
        # 1. 标记空行
        df['is_blank'] = df.isna().all(axis=1)
        # 2. 创建钢厂分组标识符（遇到空行则分组+1）
        df['group_id'] = df['is_blank'].cumsum()
        # 3. 按分组填充钢厂名称
        df['钢厂'] = df.groupby('group_id')['钢厂'].ffill()
        # 4. 过滤掉空行和不可发的记录
        df = df[~df['is_blank']]  # 删除空行
        df = df[df['是否可发'] == 1]  # 仅保留可发的规格
        
        # 处理长度为空的情况
        df['长度'] = df['长度'].fillna('').astype(str)
        
        # 按钢厂和规格型号分组
        available = defaultdict(lambda: defaultdict(list))
        for (mill, spec), group in df.groupby(['钢厂', '规格型号']):
            available[mill][spec] = group['长度'].unique().tolist()
        
        return available
    except Exception as e:
        st.error(f"加载可发规格文件失败：{str(e)}")
        return None

def load_shipment_plan(uploaded_file):
    """加载发货计划CSV"""
    try:
        df = pd.read_csv(uploaded_file)
        required_columns = ['楼号', '规格型号', '所需吨位']
        if not all(col in df.columns for col in required_columns):
            st.error(f"发货计划文件缺少必要列，请确保包含：{required_columns}")
            return None
        
        df = df[df['所需吨位'] > 0]
        df = df.dropna()
        
        plan = defaultdict(dict)
        for _, row in df.iterrows():
            building = str(row['楼号'])
            spec = row['规格型号']
            tonnage = float(row['所需吨位'])
            plan[building][spec] = tonnage
        return plan
    except Exception as e:
        st.error(f"加载发货计划文件失败：{str(e)}")
        return None

# ---------------------- 主界面 ----------------------
def main():
    st.set_page_config(page_title="钢筋发货智能体", layout="wide")
    st.title("钢筋发货智能体")
    st.markdown("### 按楼号分组，同一楼号内按规格排序，相同规格内按价差降序排序")
    
    # 初始化session_state存储设置
    if 'enable_12m_addition' not in st.session_state:
        st.session_state.enable_12m_addition = True  # 默认启用12m加价
    if 'last_enable_state' not in st.session_state:
        st.session_state.last_enable_state = st.session_state.enable_12m_addition
    if 'show_only_best' not in st.session_state:
        st.session_state.show_only_best = False  # 默认显示全部记录
    if 'last_show_state' not in st.session_state:
        st.session_state.last_show_state = st.session_state.show_only_best
    if 'tonnage_tolerance' not in st.session_state:
        st.session_state.tonnage_tolerance = 1.0  # 默认允许1吨偏差
    if 'last_tolerance_state' not in st.session_state:
        st.session_state.last_tolerance_state = st.session_state.tonnage_tolerance
    
    # 获取当前日期用于文件名
    current_date = datetime.now().strftime("%Y%m%d")
    
    with st.sidebar:
        st.header("数据输入")
        base_price_file = st.file_uploader("1. 上传每日基价CSV", type="csv")
        available_specs_file = st.file_uploader("2. 上传每日可发规格CSV", type="csv")
        shipment_plan_file = st.file_uploader("3. 上传发货计划CSV", type="csv")
        
        st.markdown("---")
        st.header("加价规则设置")
        st.checkbox(
            "启用特定钢厂12m规格加价",
            value=st.session_state.enable_12m_addition,
            key="enable_12m_addition",
            help="对中新、徐钢、河南闽源的12m螺纹钢加价30元/吨"
        )
        
        with st.expander("加价规则详情"):
            st.markdown("""
            **适用条件**：
            - **钢厂**：中新、徐钢、河南闽源
            - **长度**：12m
            - **规格**：HRB400E12, HRB400E14, HRB400E16, HRB400E18, 
                       HRB400E20, HRB400E22, HRB400E25
            - **加价金额**：30元/吨
            
            **说明**：取消勾选将禁用上述加价规则
            """)
        
        st.markdown("---")
        st.header("发货设置")
        tonnage_tolerance = st.slider(
            "允许吨位偏差范围(吨)",
            min_value=0.5,
            max_value=2.0,
            value=st.session_state.tonnage_tolerance,
            step=0.1,
            key="tonnage_tolerance",
            help="实际发货吨位与计划吨位的最大允许偏差，超出则增加一件"
        )
        
        st.markdown("---")
        st.header("显示设置")
        st.checkbox(
            "仅显示最高价差记录",
            value=st.session_state.show_only_best,
            key="show_only_best",
            help="勾选后表格将只显示每个规格中价差最高的记录"
        )
        
        st.markdown("---")
        st.subheader("排序说明")
        st.info("""
        当前排序方式：
        1. 规格汇总行（楼号为空）排在最前面
        2. 按**楼号**升序排列
        3. 同一楼号内按**规格**升序排列（按规格中的数字大小）
        4. 相同规格内按**价差**降序排列
        """)
    
    # 检测设置变更并提示用户刷新
    if (st.session_state.last_enable_state != st.session_state.enable_12m_addition or 
        st.session_state.tonnage_tolerance != st.session_state.last_tolerance_state):
        st.session_state.last_enable_state = st.session_state.enable_12m_addition
        st.session_state.last_tolerance_state = st.session_state.tonnage_tolerance
        st.warning("⚠️ 加价规则或发货设置已更改，请重新上传文件或点击下方按钮刷新结果")
        if st.button("刷新结果"):
            st.experimental_rerun()
    
    if base_price_file and available_specs_file and shipment_plan_file:
        with st.spinner("正在生成人工发货价差表..."):
            # 加载基础数据
            weight_dict = load_steel_weights("钢厂件重数据.csv")
            pricing_rules = load_pricing_rules("pricing_rules.json")
            
            # 加载用户输入数据
            base_prices_dict = load_daily_base_prices(base_price_file)
            available_specs = load_available_specs(available_specs_file)
            shipment_plan = load_shipment_plan(shipment_plan_file)
            
            if None in [weight_dict, pricing_rules, base_prices_dict, available_specs, shipment_plan]:
                st.error("数据加载失败，请检查输入文件格式")
                return
            
            # 生成人工发货价差表，传递加价规则设置和吨位偏差参数
            manual_table = generate_manual_pricing_table(
                shipment_plan, available_specs, weight_dict, pricing_rules, base_prices_dict,
                enable_12m_addition=st.session_state.enable_12m_addition,
                tonnage_tolerance=st.session_state.tonnage_tolerance
            )
            
            if manual_table.empty:
                st.warning("没有生成价差表数据，请检查输入文件")
                return
            
            # 筛选出有效记录（排除无库存和无有效价差）
            valid_table = manual_table[~manual_table['is_out_of_stock']]
            
            # 筛选出每个规格中价差最高的记录用于汇总统计
            if not valid_table.empty:
                idx = valid_table.groupby(['楼号', '规格'])['价差（元/吨）'].idxmax()
                summary_table = valid_table.loc[idx]
                
                # 计算计划与发货总量统计（仅基于价差最高的有效记录）
                plan_total_weight = summary_table['计划吨位'].sum()
                ship_total_pieces = summary_table['发货件数'].sum()
                ship_total_weight = summary_table['发货吨位'].sum()
                
                # 计算总利润（使用新公式：发货吨位 × 价差）
                total_profit = (summary_table['发货吨位'] * summary_table['价差（元/吨）']).sum().round(2)
                
                # 计算平均利润：(各楼号各规格最高价差×发货吨位之和)÷最高价差发货总吨位
                weighted_sum = (summary_table['价差（元/吨）'] * summary_table['发货吨位']).sum()
                avg_profit = weighted_sum / ship_total_weight if ship_total_weight > 0 else 0
                avg_profit = round(avg_profit, 2)
            else:
                plan_total_weight = 0
                ship_total_pieces = 0
                ship_total_weight = 0
                total_profit = 0
                avg_profit = 0
            
            # 显示汇总统计（增加平均利润）
            st.subheader("### 计划与发货数量汇总（仅含最优价差记录）")
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("计划总吨位", f"{plan_total_weight:.2f}吨")
            with col2:
                st.metric("发货总件数", f"{ship_total_pieces}件")
            with col3:
                st.metric("发货总吨位", f"{ship_total_weight:.3f}吨")  # 显示3位小数
            with col4:
                st.metric("总利润", f"¥{total_profit:,.2f}")
            with col5:
                st.metric("平均利润", f"¥{avg_profit:.2f}/吨")  # 新增平均利润指标
            
            # 显示无库存规格警告
            out_of_stock_count = manual_table['is_out_of_stock'].sum() if not manual_table.empty else 0
            if out_of_stock_count > 0:
                st.warning(f"注意：有 {out_of_stock_count} 个规格无库存或无有效价差")
            
            # 根据用户选择决定是否只显示最高价差记录
            if st.session_state.show_only_best and not manual_table.empty:
                # 分离有效记录和无库存记录
                valid_records = manual_table[~manual_table['is_out_of_stock']]
                out_of_stock_records = manual_table[manual_table['is_out_of_stock']]
                
                # 只保留每个规格中价差最高的有效记录
                if not valid_records.empty:
                    best_idx = valid_records.groupby(['楼号', '规格'])['价差（元/吨）'].idxmax()
                    best_records = manual_table.loc[best_idx]
                else:
                    best_records = pd.DataFrame()
                
                # 合并最高价差记录和无库存记录
                display_table = pd.concat([best_records, out_of_stock_records])
                
                # 重新排序以保持一致性
                display_table = display_table.sort_values(
                    ['规格排序', '楼号', '价差（元/吨）'], 
                    ascending=[True, True, False]
                )
            else:
                display_table = manual_table
            
            # 显示表格
            st.subheader("### 人工发货决策价差表")
            styled_table = format_manual_table(display_table)
            st.dataframe(styled_table, use_container_width=True)
            
            # 准备导出数据
            export_df = display_table.copy()
            
            # 下载功能 - Excel（带规格汇总行高亮）
            excel_data = format_excel_with_highlight(export_df)
            st.download_button(
                label="📊 下载Excel结果",
                data=excel_data,
                file_name=f"{current_date}_发货依据.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download-excel"
            )
            
            # 下载功能 - CSV
            if st.session_state.show_only_best and not manual_table.empty:
                download_df = display_table.drop(columns=['规格排序', 'is_max_diff', 'is_out_of_stock', 'is_summary'])
            else:
                download_df = manual_table.drop(columns=['规格排序', 'is_max_diff', 'is_out_of_stock', 'is_summary'])
            
            csv = download_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📄 下载CSV结果",
                data=csv,
                file_name=f"{current_date}_发货依据.csv",
                mime="text/csv",
                key="download-csv"
            )
            
            # 显示统计信息
            if not valid_table.empty:
                st.success(f"所有可发规格总利润：¥{total_profit:,.2f}，平均利润：¥{avg_profit:.2f}/吨")
            else:
                st.warning("没有可发规格的有效价差记录，无法计算总利润和平均利润")
            
    else:
        st.info("请上传所有必要的CSV文件（每日基价、可发规格、发货计划）")
        with st.expander("查看文件格式要求及功能说明"):
            st.markdown("""
            ### 功能亮点
            1. **规格汇总行**：Excel文件最上方显示每日可发规格的最高价差记录（楼号为空）
            2. **高亮显示**：规格汇总行自动以黄色高亮
            3. **发货吨位**：已调整为保留3位小数
            4. **利润计算**：总利润 = 发货吨位 × 价差
            5. **新增指标**：汇总统计中增加平均利润（加权平均）
            6. **双格式下载**：同时支持Excel和CSV格式导出
            
            ### 每日可发规格CSV格式示例
            ```csv
            钢厂,规格型号,长度,是否可发
            中新,HRB400E12,9m,0  # 自动过滤
            中新,HRB400E12,12m,1  # 正常加载
            中新,HRB400E6,,1  # 盘螺空长度
            ...
            ```
            """)
    
    # 检测显示设置变更并提示
    if st.session_state.last_show_state != st.session_state.show_only_best:
        st.session_state.last_show_state = st.session_state.show_only_best
        st.success("表格显示已更新为" + ("仅最高价差记录" if st.session_state.show_only_best else "全部记录"))

if __name__ == "__main__":
    main()
